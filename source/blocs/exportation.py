import os
import subprocess
import platform
import tkinter as tk
from tkinter import filedialog

import unidecode
import reportlab as rl
from reportlab.lib import units, pagesizes, styles
from reportlab.pdfgen import canvas
from reportlab.platypus import flowables
import openpyxl

from . import config


cm = rl.lib.units.cm


def filenamify(str):
    return "".join(c for c in unidecode.unidecode(str) if c not in r'.\/:*?"<>|')       # str -> ACSII et enlève caractères interdits dans les noms de fichiers


def open_file(filepath):
    """Provoque l'ouverture du fichier dans <filepath> par l'application par défaut de l'OS hôte

    Si <filepath> pointe vers un dossier, l'ouvre dans l'explorateur de fichiers
    (https://stackoverflow.com/a/435669)
    """
    if platform.system() == "Windows":      # Windows
        os.startfile(filepath)
    elif platform.system() == "Darwin":     # macOS
        subprocess.call(("open", filepath))
    else:                                   # linux variants
        subprocess.run(["xdg-open", filepath], check=True)


def exporter_fiches_eleves():
    lien_dossier = tk.filedialog.askdirectory(title="Sélectionnez un dossier pour exporter les fiches PDF individuelles :")
    if lien_dossier:        # Dossier sélectionné (pas Annuler)
        with config.ContextPopup(config.root, "Exportation...") as popup:
            N_pdf = len(config.clients)
            for i_pdf, client in enumerate(config.clients):
                pdf_client(client, lien_dossier, open_pdf=False, i_pdf=i_pdf, N_pdf=N_pdf, popup=popup)

        if tk.messagebox.askyesno(title="Exportation", message=f"Fiches élèves exportées dans {lien_dossier} !\n\nL'ouvrir dans l'explorateur de fichiers ?"):
            open_file(lien_dossier)


def exporter_fiches_spectacles():
    lien_dossier = tk.filedialog.askdirectory(title="Sélectionnez un dossier où les fiches spectacles seront exportées :")
    if lien_dossier:        # Dossier sélectionné (pas Annuler)
        with config.ContextPopup(config.root, "Exportation...") as popup:
            for spec in config.spectacles:
                pdf_spectacle(spec, lien_dossier, open_pdf=False, popup=popup)

        if tk.messagebox.askyesno(title="Exportation", message=f"Fiches spectacles exportées dans {lien_dossier} !\n\nL'ouvrir dans l'explorateur de fichiers ?"):
            open_file(lien_dossier)



def pdf_client(client, lien_dossier=None, open_pdf=True, i_pdf=None, N_pdf=None, popup=None):

    if not lien_dossier:
        lien_dossier = tk.filedialog.askdirectory(title=f"Sélectionnez un dossier pour exporter la fiche PDF de {client.prenom} {client.nom} :")

    filepath = os.path.join(os.path.normpath(lien_dossier), f"{filenamify(client.nomprenom)}.pdf")

    etape = f" {i_pdf+1}/{N_pdf}" if N_pdf else "..."
    with config.ContextPopup(config.root, f"Génération du PDF{etape}", existing=popup):
        canvas = rl.pdfgen.canvas.Canvas(filepath, pagesize=rl.lib.pagesizes.A4)
        canvas.setAuthor("Club Q ESPCI Paris - PSL")
        canvas.setTitle(f"Compte-rendu {client.prenom}{client.nom}")
        canvas.setSubject(f"Saison {config.saison.nom}")

        styles = rl.lib.styles.getSampleStyleSheet()
        styleN = styles["Normal"]
        canvas.setFont("Times-Bold", 18)


        descr = [
            f"Nom : <b>{client.nom}</b>",
            f"Prénom : <b>{client.prenom}</b>",
            f"Promo : <b>{client.promo or client.autre}</b>",
            f"Adresse e-mail : <b>{client.email}</b>"
        ]


        specs = [f"{attrib.spectacle.nom} - {attrib.places_attribuees} place(s) -\t\t {attrib.spectacle.unit_price} € /place"
                 for attrib in client.attributions()]

        specs.append("-"*158)
        specs.append(f"Total à payer : {client.a_payer} €.")


        # Conversion en objets reportlab
        blocs_descr = []
        for txt in descr:
            blocs_descr.append(rl.platypus.Paragraph(txt, styleN))
            blocs_descr.append(rl.platypus.Spacer(1, 0.2*cm))

        blocs_specs = []
        for txt in specs:
            blocs_specs.append(rl.platypus.Paragraph(txt, styleN))
            blocs_specs.append(rl.platypus.Spacer(1, 0.2*cm))


        # IMPRESSION
        canvas.setFont("Times-Bold", 18)
        canvas.drawString(1*cm, 28*cm, f"Compte-rendu Club Q")

        canvas.setFont("Times-Bold", 12)
        canvas.drawString(1*cm, 27.2*cm, f"Saison {config.saison.nom}")

        dX, dY = canvas.drawImage(config.lien_logo_pc, 14*cm, 27.22*cm, width=5.63*cm, height=1.2*cm, mask="auto")
        dX, dY = canvas.drawImage(config.lien_logo_q, 11.5*cm, 26.8*cm, width=2*cm, height=2*cm, mask="auto")

        frame_descr = rl.platypus.Frame(1*cm, 23.5*cm, 19*cm, 3*cm, showBoundary=True)
        frame_descr.addFromList(blocs_descr, canvas)

        frame_specs = rl.platypus.Frame(1*cm, 1*cm, 19*cm, 22*cm, showBoundary=True)
        frame_specs.addFromList(blocs_specs, canvas)

        canvas.showPage()


        # SAUVEGARDE
        canvas.save()

        if open_pdf and tk.messagebox.askyesno(title="Exportation", message=f"Fiche élève individuelle exportée !\n\nL'ouvrir ?"):
            open_file(filepath)

    return filepath




def pdf_spectacle(spec, lien_dossier=None, open_pdf=True, i_pdf=None, N_pdf=None, popup=None):

    if not lien_dossier:
        lien_dossier = tk.filedialog.askdirectory(title=f"Sélectionnez un dossier pour exporter la fiche PDF de {spec.nom} :")

    filepath = os.path.join(os.path.normpath(lien_dossier), f"{filenamify(spec.nom)}.pdf")

    etape = f" {i_pdf+1}/{N_pdf}" if N_pdf else "..."
    with config.ContextPopup(config.root, f"Génération du PDF{etape}", existing=popup):
        canvas = rl.pdfgen.canvas.Canvas(filepath, pagesize=rl.lib.pagesizes.A4)
        styles = rl.lib.styles.getSampleStyleSheet()
        styleN = styles["Normal"]
        canvas.setFont("Times-Bold", 18)


        descr = [
            f"Titre : <b>{spec.nom}</b> ({spec.categorie or 'catégorie inconnue'})",
            f"Salle : <b>{spec.lieu()}</b>, {spec.salle.adresse or 'adresse inconnue'}",
            f"Date : <b>{spec.date or 'inconnue'}  {spec.heure or '(heure inconnue)'}</b>",
            # f"Heure : <b>{spec.heure}</b>",
            f"Nombre de places : <b>{spec.nb_tickets}</b> – Demandées : <b>{spec.nb_places_demandees()}</b> – Attribuées : <b>{spec.nb_places_attribuees()}</b>",
        ]

        eleves = [f"{attrib.client.nomprenom}, {attrib.client.promo or attrib.client.autre}, {attrib.client.email} - {attrib.places_attribuees} place(s)"
                  for attrib in spec.attributions()]


        # Conversion en objets reportlab

        blocs_descr = []
        for txt in descr:
            blocs_descr.append(rl.platypus.Paragraph(txt, styleN))
            blocs_descr.append(rl.platypus.Spacer(1, .2*cm))

        blocs_eleves = []
        for txt in eleves:
            blocs_eleves.append(rl.platypus.Paragraph(txt, styleN))
            blocs_eleves.append(rl.platypus.Spacer(1, .2*cm))


        # IMPRESSION PAGE 1
        canvas.setFont("Times-Bold", 18)
        canvas.drawString(1*cm, 28*cm, f"Compte-rendu {spec.nom}")

        canvas.setFont("Times-Bold", 12)
        canvas.drawString(1*cm, 27.2*cm, f"Saison {config.saison.nom}")

        dX, dY = canvas.drawImage(config.lien_logo_pc, 14*cm, 27.22*cm, width=5.63*cm, height=1.2*cm, mask="auto")
        dX, dY = canvas.drawImage(config.lien_logo_q, 11.5*cm, 26.8*cm, width=2*cm, height=2*cm, mask="auto")

        frame_descr = rl.platypus.Frame(1*cm, 23.5*cm, 19*cm, 3*cm, showBoundary =1)
        frame_descr.addFromList(blocs_descr, canvas)

        frame_eleves = rl.platypus.Frame(1*cm, 1*cm, 19*cm, 22*cm, showBoundary =1)
        frame_eleves.addFromList(blocs_eleves, canvas)

        canvas.showPage()


        # PAGE 2 (liste d'attente)

        voeux = [voeu for voeu in spec.voeux() if not voeu.places_attribuees]

        if voeux:
            voeux.sort(key=lambda a: a.client.mecontentement or 0, reverse=True)

            voeux_txt = ["Liste d'attente :"]
            for voeu in voeux:
                voeux_txt.append(f"M: {round(voeu.client.mecontentement or 0, 2)} - {voeu.client.nomprenom} - {voeu.places_demandees} place(s) ({voeu.places_minimum} min) - Voeu n°{voeu.priorite}")

        else:
            voeux_txt = ["Pas de liste d'attente."]


        blocs_descr = []        # Détruits à la précédente impression...
        for txt in descr:
            blocs_descr.append(rl.platypus.Paragraph(txt, styleN))
            blocs_descr.append(rl.platypus.Spacer(1, .2*cm))

        blocs_voeux = []
        for txt in voeux_txt:
            blocs_voeux.append(rl.platypus.Paragraph(txt, styleN))
            blocs_voeux.append(rl.platypus.Spacer(1, .2*cm))


        # IMPRESSION PAGE 2
        canvas.setFont("Times-Bold", 18)
        canvas.drawString(1*cm, 28*cm, f"Liste d'attente {spec.nom}")

        canvas.setFont("Times-Bold", 12)
        canvas.drawString(1*cm, 27.2*cm, f"Saison {config.saison.nom}")

        dX, dY = canvas.drawImage(config.lien_logo_pc, 14*cm, 27.22*cm, width=5.63*cm, height=1.2*cm, mask="auto")
        dX, dY = canvas.drawImage(config.lien_logo_q, 11.5*cm, 26.8*cm, width=2*cm, height=2*cm, mask="auto")

        frame_descr = rl.platypus.Frame(1*cm, 23.5*cm, 19*cm, 3*cm, showBoundary =1)
        frame_descr.addFromList(blocs_descr, canvas)

        frame_eleves = rl.platypus.Frame(1*cm, 1*cm, 19*cm, 22*cm, showBoundary =1)
        frame_eleves.addFromList(blocs_voeux, canvas)

        canvas.showPage()


        # SAUVEGARDE
        canvas.save()

        if open_pdf and tk.messagebox.askyesno(title="Exportation", message=f"Fiche spectacle exportée !\n\nL'ouvrir ?"):
            open_file(filepath)

    return filepath



def exporter_excel_prix():
    with config.ContextPopup(config.root, "Génération..."):

        # Code assez (très) moche, attention - il y a moyen de faire ça beaucoup mieux, mais bon...
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Récapitulatif"

        sheet["A1"] = f"Club Q – Récapitulatif saison {config.saison.nom}"
        sheet["A1"].font = openpyxl.styles.Font(bold=True)

        sheet["A2"] = "Début :"
        sheet["B2"] = config.saison.debut
        sheet["B2"].number_format = "dd/mm/yyyy"
        sheet["D2"] = "Fin :"
        sheet["E2"] = config.saison.fin
        sheet["E2"].number_format = "dd/mm/yyyy"

        sheet["A3"] = "Spectacles :"
        sheet["B3"] = len(config.spectacles)
        sheet["D3"] = "Inscrits :"
        sheet["E3"] = len(config.clients)

        sheet["A4"] = "Places proposées :"
        sheet["B4"] = sum((spec.nb_tickets or 0) for spec in config.spectacles)
        sheet["D4"] = "Places vendues :"
        sheet["E4"] = sum((voeu.places_attribuees or 0) for voeu in config.voeux)

        sheet["A6"] = "NOM"
        sheet["B6"] = "Prénom"
        sheet["C6"] = "Promo/autre"
        sheet["D6"] = "Adresse mail"
        sheet["E6"] = "Places"
        sheet["F6"] = "Somme due"
        sheet["G6"] = "Payé ?"

        npa = 0
        tot = 0
        for i, client in enumerate(sorted(config.clients, key=lambda c: c.nomprenom)):
            sheet.cell(row=7 + i, column=1).value = client.nom.upper()
            sheet.cell(row=7 + i, column=2).value = client.prenom.title()
            sheet.cell(row=7 + i, column=3).value = client.promo or client.autre

            sheet.cell(row=7 + i, column=4).value = client.email
            if client.email and "@" in client.email:
                sheet.cell(row=7 + i, column=4).hyperlink = f"mailto:{client.email}"
                sheet.cell(row=7 + i, column=4).style = "Hyperlink"

            placesattr = client.nb_places_attribuees()
            sheet.cell(row=7 + i, column=5).value = placesattr
            npa += placesattr
            apayer = client.a_payer
            sheet.cell(row=7 + i, column=6).value = apayer
            sheet.cell(row=7 + i, column=6).number_format = "#,##0.00 €_-"
            tot += apayer


        # sheet.auto_filter.ref = f"A6:G{7 + i}"
        tab = openpyxl.worksheet.table.Table(displayName="Données", ref=f"A6:G{7 + i}")

        # Add a default style with striped rows and banded columns
        style = openpyxl.worksheet.table.TableStyleInfo(
            name="TableStyleLight1", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False
            )
        tab.tableStyleInfo = style
        sheet.add_table(tab)

        sheet.cell(row=7 + i + 2, column=1).value = "Total"
        sheet.cell(row=7 + i + 2, column=1).font = openpyxl.styles.Font(bold=True)

        sheet.cell(row=7 + i + 2, column=5).value = npa
        sheet.cell(row=7 + i + 2, column=5).font = openpyxl.styles.Font(bold=True)

        sheet.cell(row=7 + i + 2, column=6).value = tot
        sheet.cell(row=7 + i + 2, column=6).font = openpyxl.styles.Font(bold=True)
        sheet.cell(row=7 + i + 2, column=6).number_format = "#,##0.00 €_-"


        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 20
        sheet.column_dimensions["C"].width = 14
        sheet.column_dimensions["D"].width = 35
        sheet.column_dimensions["E"].width = 12
        sheet.column_dimensions["F"].width = 14
        sheet.column_dimensions["G"].width = 12

    filepath = tk.filedialog.asksaveasfilename(title="Enregistrement du PDF récapitulatif des sommes à payer", initialfile=filenamify(f"Récap Club Q S{config.saison.nom}") + ".xlsx", filetypes=[("Fichiers Excel", "*.xlsx"), ("Tous les fichiers", "*")], defaultextension=".xlsx")
    if filepath:    # Pas Annuler
        workbook.save(filepath)
        if tk.messagebox.askyesno(title="Fichier exporté", message="Excel récapitulatif exporté ! L'ouvrir ?"):
            open_file(filepath)
